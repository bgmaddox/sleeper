#!/usr/bin/env python3
"""
scripts/parse_sidebet_xlsx.py

One-shot parser: reads historical side bet Excel files (2019–2023) and outputs
Python dict literals ready to paste into SIDE_BET_SEASONS in sleeper_core.py.

2024 is hardcoded directly — the Excel file for that year is a stale copy of
2023 with no real data; winners were confirmed from league records.

Three phases:
  1. Discovery pass — prints all unique raw winner tokens by year; halts if any
     are unmapped in DISPLAY_NAME_MAP.
  2. Duplicate detection — MD5-hashes each Challenges sheet (cols A+B only);
     halts if two years collide unexpectedly.
  3. Parse pass — outputs formatted Python dict literals.

Usage:
    cd "Sleeper Project"
    source .venv/bin/activate
    python scripts/parse_sidebet_xlsx.py

Committed to scripts/ as permanent documentation of how SIDE_BET_SEASONS was
populated. Not imported by the app.
"""

import hashlib
import re
import sys

import openpyxl

# ── Name → Sleeper username mapping ──────────────────────────────────────────
# Excel files use informal display names; these map them to Sleeper usernames.

DISPLAY_NAME_MAP = {
    "Brett":      "bgmaddox",
    "JT":         "JTizzzzle",
    "Jack":       "jlglover",
    "Reclam":     "RReclam",          # NOTE: 2019 uses GurlyGirls (old username); fix manually after paste
    "LB":         "RReclam",         # "LB/Reclam" in 2019 — same person (also needs GurlyGirls fix for 2019)
    "Chip":       "BMoreBallers88",
    "Rachael":    "RascalHazard",
    "Erin":       "eegrady",
    "Hunter":     "jhuntmadd",
    "Liam":       "RossLikeSauce",
    "Ross":       "RossLikeSauce",
    "SG":         "sgmaddox",
    "Stuart":     "BillyRayGonnaGetcha",  # 2021 only; same person as Billy
    "Aaron":      "YouthPastor",
    "Billy":      "BillyRayGonnaGetcha",
    "Rachel D.":  "SweetDizzzzzle",
    "Liz":        "DirtyCommie",
    "Jess":       "InfiniteJesse",
}

# ── Excel source files (2019–2023 only) ───────────────────────────────────────
# 2024 is hardcoded below — its Excel file is a stale copy of 2023.

FILE_YEAR_MAP = {
    2019: "sidebets_historical/Side Bets - Legacy League '19.xlsx",
    2020: "sidebets_historical/Side Bets - Legacy League '20.xlsx",
    2021: "sidebets_historical/Legacy League '21.xlsx",
    2022: "sidebets_historical/Legacy League '22.xlsx",
    2023: "sidebets_historical/Legacy League '23.xlsx",
}

# ── 2024 hardcoded data ────────────────────────────────────────────────────────
# Winners confirmed directly from league records (already Sleeper usernames).
# The 2024 Excel file (Legacy League '24.xlsx) is byte-for-byte identical to
# 2023 — formula caches were never refreshed — so it cannot be used as a source.

HARDCODED_2024 = {
    1:  {"name": "I'm flying, Jack!",
         "desc": "Team with the highest score (starters only)",
         "winner": "jlglover"},
    2:  {"name": "Look At These TDs",
         "desc": "Team with the most offensive touchdowns scored",
         "winner": "bgmaddox"},
    3:  {"name": "Endzones that way -->",
         "desc": "Team with the worst performing defense (active or bench)",
         "winner": "JTizzzzle"},
    4:  {"name": "Blackjack",
         "desc": "Team with a starter closest to 21 points without going over",
         "winner": "jhuntmadd"},
    5:  {"name": "The Replacements",
         "desc": "Team with the highest total points for their bench",
         "winner": "sgmaddox"},
    6:  {"name": "Like A Boss",
         "desc": "Team with the biggest margin of victory",
         "winner": "bgmaddox"},
    7:  {"name": "Campus Rush Week",
         "desc": "Total rush yards for team (active or bench)",
         "winner": "eegrady"},
    8:  {"name": "Soothsayer",
         "desc": "Players submit guesses for their total points by Thursday afternoon. Team closest to projection wins.",
         "winner": "sgmaddox"},
    9:  {"name": "Keeping it Tight",
         "desc": "Team with best performing tight end (active or bench)",
         "winner": "JTizzzzle"},
    10: {"name": "NFL Franchise Week",
         "desc": "Team with the highest point total of players from the same franchise (active or bench)",
         "winner": "RossLikeSauce"},
    11: {"name": "Please not the Jets (Trade Deadline Week)",
         "desc": "Team with the most trades this seasons wins",
         "winner": "jlglover"},
    12: {"name": "Go Long",
         "desc": "Team with the Starting QB with the highest completion % (over 10 throws)",
         "winner": "RossLikeSauce"},
    13: {"name": "Coffee's For Closers",
         "desc": "Team that beats its opponent by the smallest margin of victory",
         "winner": "RReclam"},
    14: {"name": "Breaking of the Tie (if needed)",
         "desc": "Choose 3 non-QB players. Highest combined total wins.",
         "winner": "RossLikeSauce"},
}

# ── Helpers ───────────────────────────────────────────────────────────────────

WEEK_RE  = re.compile(r'WEEK\s+(\d+)', re.IGNORECASE)
SPLIT_RE = re.compile(r'\s*/\s*')


def load_ws(path):
    return openpyxl.load_workbook(path, data_only=True)["Challenges"]


def winner_tokens(raw):
    """Split a raw winner cell into individual name tokens; exclude n/a / empty."""
    if raw is None:
        return []
    s = str(raw).strip()
    if not s or s.lower() == "n/a":
        return []
    return [t.strip() for t in SPLIT_RE.split(s) if t.strip()]


def sheet_md5(ws):
    # Hash only cols A+B so extra bookkeeping columns don't mask data duplicates.
    body = "|".join(
        f"{str(row[0].value)}:{str(row[1].value if len(row) > 1 else '')}"
        for row in ws.iter_rows()
        if row and row[0].value
    )
    return hashlib.md5(body.encode()).hexdigest()


def week_rows(ws):
    """Yield (week_num, raw_side_bet_text, winner_raw) for WEEK rows only."""
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        m = WEEK_RE.match(str(row[0]))
        if not m:
            continue
        yield int(m.group(1)), str(row[0]).strip(), row[1]


def parse_bet_text(text):
    """'WEEK N: Name - Desc' → (name, desc). Handles hyphen, en-dash, em-dash."""
    after = re.sub(r'^WEEK\s+\d+:\s*', '', text, flags=re.IGNORECASE).strip()
    for sep in (" - ", " – ", " — "):
        if sep in after:
            name, _, desc = after.partition(sep)
            return name.strip(), desc.strip()
    return after, ""


def map_winner(tokens):
    """Map display-name tokens to ' & '-joined usernames. Deduplicates aliases."""
    if not tokens:
        return ""
    seen = {}
    for t in tokens:
        seen[DISPLAY_NAME_MAP[t]] = True
    return " & ".join(seen)


# ── Phase 1: Discovery Pass ───────────────────────────────────────────────────

print("=" * 64)
print("PHASE 1: DISCOVERY PASS  (2019–2023 Excel files)")
print("=" * 64)

unmapped_by_year = {}

for year, path in FILE_YEAR_MAP.items():
    ws = load_ws(path)
    tokens = sorted({tok for _, _, raw in week_rows(ws) for tok in winner_tokens(raw)})
    bad = [t for t in tokens if t not in DISPLAY_NAME_MAP]
    unmapped_by_year[year] = bad
    status = "✗ UNMAPPED: " + str(bad) if bad else "✓"
    print(f"  {year}: {tokens}  →  {status}")

print(f"\n  2024: (hardcoded — Excel file is a stale copy of 2023)  →  ✓")

all_bad = sorted({t for v in unmapped_by_year.values() for t in v})
if all_bad:
    print()
    print("!" * 64)
    print("HALT — add these names to DISPLAY_NAME_MAP before re-running:")
    for year, names in unmapped_by_year.items():
        if names:
            print(f"  {year}: {names}")
    print("!" * 64)
    sys.exit(1)

print("\nAll names mapped. ✓")


# ── Phase 2: Duplicate Detection ─────────────────────────────────────────────

print()
print("=" * 64)
print("PHASE 2: DUPLICATE DETECTION  (2019–2023 only)")
print("=" * 64)

hashes = {}
for year, path in FILE_YEAR_MAP.items():
    h = sheet_md5(load_ws(path))
    hashes[year] = h
    print(f"  {year}: {h}")

seen_hashes = {}
collisions = []
for year, h in hashes.items():
    if h in seen_hashes:
        collisions.append((seen_hashes[h], year))
    else:
        seen_hashes[h] = year

if collisions:
    print()
    print("!" * 64)
    print("HALT — unexpected duplicate Challenges sheet content:")
    for y1, y2 in collisions:
        print(f"  {y1} and {y2} have identical content (same MD5).")
    print("Investigate before proceeding.")
    print("!" * 64)
    sys.exit(1)

print("\nNo duplicate sheets. ✓")


# ── Phase 3: Parse Pass ───────────────────────────────────────────────────────

print()
print("=" * 64)
print("PHASE 3: PARSE OUTPUT")
print("=" * 64)
print()
print("Paste the block below into SIDE_BET_SEASONS in sleeper_core.py")
print("(above the existing 2025 entry):")
print()

# 2019–2023 from Excel
for year, path in FILE_YEAR_MAP.items():
    ws = load_ws(path)
    print(f"    {year}: {{")
    for wk, text, raw in week_rows(ws):
        name, desc = parse_bet_text(text)
        winner_str = map_winner(winner_tokens(raw))
        print(
            f"        {wk:<3}: "
            f'{{\"name\": {name!r:<46},'
            f' \"desc\": {desc!r:<92},'
            f' \"winner\": {winner_str!r}}},'
        )
    print("    },")

# 2024 hardcoded
print(f"    2024: {{")
for wk, cfg in HARDCODED_2024.items():
    n, d, w = cfg["name"], cfg["desc"], cfg["winner"]
    print(
        f"        {wk:<3}: "
        f'{{\"name\": {n!r:<46},'
        f' \"desc\": {d!r:<92},'
        f' \"winner\": {w!r}}},'
    )
print("    },")
